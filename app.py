import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        price = float(sheet.cell(row, 3).value.replace('$', ''))
        sheet.cell(row, 4).value = round(price * 0.9, 2)

    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, "F2")

    wb.save(filename)
